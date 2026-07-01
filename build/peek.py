import openpyxl

root = r'E:\WorkSpace\Competition\项目\燕理智能招生\资料'

wb = openpyxl.load_workbook(root + '\\专业数据.xlsx', data_only=True)
ws = wb.active
print('Sheet:', wb.sheetnames[0])
print('Rows:', ws.max_row, 'Cols:', ws.max_column)
print()
for row in ws.iter_rows(min_row=1, max_row=5, values_only=True):
    print(list(row))
print('...')
for row in ws.iter_rows(min_row=ws.max_row-2, max_row=ws.max_row, values_only=True):
    print(list(row))
